"""

Library use:
    hashlib----Algorithms
    openpyxl---- Excel

"""

import hashlib
import os
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import messagebox


file_path="C:/Users/ADMIN/python files/Major_Projects/User.xlsx"

def encrypt_password(password):
    encrypted_password=hashlib.sha256(password.encode()).hexdigest()

    return encrypted_password


def create_file():
    if not os.path.exists(file_path):
        wb=Workbook()
        sheet=wb.active
        sheet.append(["Username","Password"])
        wb.save(file_path)

create_file()

def register_user():
    username=entry_username.get()
    password=entry_password.get()

    if username=="" or password=="":
        print("All field are required")
        messagebox.showerror("Error","All field are required")
        return

    encrypted_password=encrypt_password(password)

    wb=load_workbook(file_path)
    sheet=wb.active


    for row in sheet.iter_rows(min_row=2,values_only=True):
        if row[0]==username:
            print("Username already exists")
            messagebox.showerror("Error","Username already exist")
            clean()
            return
    sheet.append([username,encrypted_password])

    wb.save(file_path)
    print("User Registered Successfully")
    messagebox.showinfo("Success","Username Registered Successfully")
    clean()
    




def login_user():
    username=entry_username.get()
    password=entry_password.get()

    encrypted_password=encrypt_password(password)

    wb=load_workbook(file_path)
    sheet=wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if username==row[0] and encrypted_password==row[1]:
            print("Login successfully")
            messagebox.showinfo("Success","Login Successfully")
            clean()
            return
    print("Login unsuccessful")   

    messagebox.showerror("Error","Login unsuccessful")
    clean()


def  clean():
    entry_username.delete(0,END)
    entry_password.delete(0,END)
    


root=Tk()
root.title("Secure login System")
root.geometry("700x500")
root.resizable(False,False)

Label(root,width=25,text="Username").pack(pady=10)

entry_username=Entry(root)
entry_username.pack()


Label(root,width=25,text="Password").pack(pady=10)

entry_password=Entry(root,show="*")
entry_password.pack()

Button(root,text="Register",width=20,command=register_user).pack(pady=10)
Button(root,text="Login",width=20,command=login_user).pack(pady=10)


root.mainloop()